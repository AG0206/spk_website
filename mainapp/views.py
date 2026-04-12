import io
from functools import wraps

from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.core.paginator import Paginator
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.dateparse import parse_date

from .models import Contact, Event, Gallery


# ─────────────────────────────────────────────────────────────
# DECORATOR — require staff login
# ─────────────────────────────────────────────────────────────
def admin_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or not request.user.is_staff:
            return redirect('custom_admin_login')
        return view_func(request, *args, **kwargs)
    return wrapper


def _admin_context(request):
    """Inject sidebar counts into every admin view."""
    return {
        'event_count':   Event.objects.filter(is_active=True).count(),
        'gallery_count': Gallery.objects.filter(is_active=True).count(),
        'unread_count':  Contact.objects.filter(is_read=False).count(),
        'total_contacts': Contact.objects.count(),
    }


# ─────────────────────────────────────────────────────────────
# AUTH
# ─────────────────────────────────────────────────────────────
def custom_admin_login(request):
    if request.user.is_authenticated and request.user.is_staff:
        return redirect('custom_admin_dashboard')

    error = None
    if request.method == 'POST':
        username = request.POST.get('username', '').strip()
        password = request.POST.get('password', '')
        user = authenticate(request, username=username, password=password)
        if user and user.is_staff:
            login(request, user)
            return redirect('custom_admin_dashboard')
        else:
            error = 'Invalid credentials or insufficient permissions.'

    return render(request, 'custom_admin/login.html', {'error': error})


def custom_admin_logout(request):
    logout(request)
    return redirect('custom_admin_login')


# ─────────────────────────────────────────────────────────────
# DASHBOARD
# ─────────────────────────────────────────────────────────────
@admin_required
def custom_admin_dashboard(request):
    ctx = _admin_context(request)
    ctx['active_nav']       = 'dashboard'
    ctx['recent_contacts']  = Contact.objects.order_by('-created_at')[:8]
    return render(request, 'custom_admin/dashboard.html', ctx)


# ─────────────────────────────────────────────────────────────
# EVENTS
# ─────────────────────────────────────────────────────────────
@admin_required
def custom_admin_events(request):
    ctx = _admin_context(request)
    ctx['active_nav'] = 'events'
    ctx['events']     = Event.objects.all().order_by('order', '-created_at')
    return render(request, 'custom_admin/events_list.html', ctx)


@admin_required
def custom_admin_event_add(request):
    ctx = _admin_context(request)
    ctx['active_nav']    = 'events'
    ctx['badge_choices'] = Event.BADGE_CHOICES
    ctx['event']         = None

    if request.method == 'POST':
        title       = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        if not title or not description:
            messages.error(request, 'Title and description are required.')
            return render(request, 'custom_admin/event_form.html', ctx)

        event = Event(
            title       = title,
            description = description,
            date_label  = request.POST.get('date_label', 'Ongoing 2025').strip(),
            location    = request.POST.get('location', 'Nagpur, MH').strip(),
            badge       = request.POST.get('badge', 'Free'),
            link_label  = request.POST.get('link_label', 'Register Interest').strip(),
            order       = int(request.POST.get('order', 0) or 0),
            is_featured = 'is_featured' in request.POST,
            is_active   = 'is_active'   in request.POST,
        )
        if request.FILES.get('image'):
            event.image = request.FILES['image']
        event.save()
        messages.success(request, f'Event "{event.title}" created successfully.')
        return redirect('custom_admin_events')

    return render(request, 'custom_admin/event_form.html', ctx)


@admin_required
def custom_admin_event_edit(request, pk):
    event = get_object_or_404(Event, pk=pk)
    ctx   = _admin_context(request)
    ctx['active_nav']    = 'events'
    ctx['badge_choices'] = Event.BADGE_CHOICES
    ctx['event']         = event

    if request.method == 'POST':
        title       = request.POST.get('title', '').strip()
        description = request.POST.get('description', '').strip()
        if not title or not description:
            messages.error(request, 'Title and description are required.')
            return render(request, 'custom_admin/event_form.html', ctx)

        event.title       = title
        event.description = description
        event.date_label  = request.POST.get('date_label', event.date_label).strip()
        event.location    = request.POST.get('location', event.location).strip()
        event.badge       = request.POST.get('badge', event.badge)
        event.link_label  = request.POST.get('link_label', event.link_label).strip()
        event.order       = int(request.POST.get('order', event.order) or 0)
        event.is_featured = 'is_featured' in request.POST
        event.is_active   = 'is_active'   in request.POST
        if request.FILES.get('image'):
            event.image = request.FILES['image']
        event.save()
        messages.success(request, f'Event "{event.title}" updated successfully.')
        return redirect('custom_admin_events')

    return render(request, 'custom_admin/event_form.html', ctx)


@admin_required
def custom_admin_event_delete(request, pk):
    event = get_object_or_404(Event, pk=pk)
    name  = event.title
    event.delete()
    messages.success(request, f'Event "{name}" deleted.')
    return redirect('custom_admin_events')


# ─────────────────────────────────────────────────────────────
# GALLERY
# ─────────────────────────────────────────────────────────────
@admin_required
def custom_admin_gallery(request):
    ctx = _admin_context(request)
    ctx['active_nav'] = 'gallery'
    ctx['gallery']    = Gallery.objects.all().order_by('order', '-created_at')

    if request.method == 'POST' and request.POST.get('action') == 'upload':
        title = request.POST.get('title', '').strip()
        image = request.FILES.get('image')
        if not title or not image:
            messages.error(request, 'Title and image file are required.')
        else:
            Gallery.objects.create(
                title    = title,
                image    = image,
                category = request.POST.get('category', 'events'),
                size     = request.POST.get('size', 'normal'),
                order    = int(request.POST.get('order', 0) or 0),
                is_active = True,
            )
            messages.success(request, f'Image "{title}" uploaded successfully.')
            return redirect('custom_admin_gallery')

    return render(request, 'custom_admin/gallery_manage.html', ctx)


@admin_required
def custom_admin_gallery_delete(request, pk):
    item = get_object_or_404(Gallery, pk=pk)
    name = item.title
    # Optionally delete file from disk too
    if item.image:
        try:
            item.image.delete(save=False)
        except Exception:
            pass
    item.delete()
    messages.success(request, f'Image "{name}" deleted.')
    return redirect('custom_admin_gallery')


# ─────────────────────────────────────────────────────────────
# CONTACTS
# ─────────────────────────────────────────────────────────────
@admin_required
def custom_admin_contacts(request):
    ctx = _admin_context(request)
    ctx['active_nav'] = 'contacts'

    qs = Contact.objects.all().order_by('-created_at')

    # Filters
    q         = request.GET.get('q', '').strip()
    status    = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()

    if q:
        from django.db.models import Q
        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) |
                       Q(subject__icontains=q) | Q(message__icontains=q))
    if status == 'unread':
        qs = qs.filter(is_read=False)
    elif status == 'read':
        qs = qs.filter(is_read=True)

    if date_from:
        d = parse_date(date_from)
        if d:
            qs = qs.filter(created_at__date__gte=d)
    if date_to:
        d = parse_date(date_to)
        if d:
            qs = qs.filter(created_at__date__lte=d)

    # Paginate
    paginator = Paginator(qs, 20)
    page_num  = request.GET.get('page', 1)
    contacts  = paginator.get_page(page_num)

    ctx['contacts']     = contacts
    ctx['total_count']  = qs.count()
    return render(request, 'custom_admin/contacts.html', ctx)


@admin_required
def custom_admin_contact_toggle_read(request, pk):
    contact      = get_object_or_404(Contact, pk=pk)
    contact.is_read = not contact.is_read
    contact.save()
    next_url = request.GET.get('next', request.META.get('HTTP_REFERER', '/adminpanel/contacts/'))
    return redirect(next_url)


@admin_required
def custom_admin_contact_delete(request, pk):
    contact = get_object_or_404(Contact, pk=pk)
    name    = contact.name
    contact.delete()
    messages.success(request, f'Message from "{name}" deleted.')
    next_url = request.GET.get('next', '/adminpanel/contacts/')
    return redirect(next_url)


# ─────────────────────────────────────────────────────────────
# EXPORT CONTACTS → EXCEL
# ─────────────────────────────────────────────────────────────
@admin_required
def custom_admin_contacts_export(request):
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl is not installed. Run: pip install openpyxl')
        return redirect('custom_admin_contacts')

    # Apply same filters as list view
    qs = Contact.objects.all().order_by('-created_at')
    q         = request.GET.get('q', '').strip()
    status    = request.GET.get('status', '').strip()
    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()

    if q:
        from django.db.models import Q
        qs = qs.filter(Q(name__icontains=q)|Q(email__icontains=q)|
                       Q(subject__icontains=q)|Q(message__icontains=q))
    if status == 'unread':
        qs = qs.filter(is_read=False)
    elif status == 'read':
        qs = qs.filter(is_read=True)
    if date_from:
        d = parse_date(date_from)
        if d: qs = qs.filter(created_at__date__gte=d)
    if date_to:
        d = parse_date(date_to)
        if d: qs = qs.filter(created_at__date__lte=d)

    # Build workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contact Messages'

    # ── Header styling ──
    header_fill = PatternFill(start_color='1C2B4A', end_color='1C2B4A', fill_type='solid')
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    gold_fill   = PatternFill(start_color='B8935A', end_color='B8935A', fill_type='solid')

    # Title row
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = 'SPK Organisation — Contact Form Submissions'
    title_cell.font  = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
    title_cell.fill  = header_fill
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # Column headers
    headers = ['#', 'Name', 'Email', 'Phone', 'Subject', 'Message', 'Date', 'Status']
    ws.append(headers)
    for col_idx, _ in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_idx)
        cell.fill      = gold_fill
        cell.font      = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 18

    # Data rows
    for i, c in enumerate(qs, 1):
        row = [
            i,
            c.name,
            c.email,
            c.phone or '',
            c.subject or '',
            c.message,
            c.created_at.strftime('%d %b %Y %H:%M'),
            'Read' if c.is_read else 'Unread',
        ]
        ws.append(row)
        r = ws.max_row
        # Alternate row shading
        if i % 2 == 0:
            light = PatternFill(start_color='E8F5FB', end_color='E8F5FB', fill_type='solid')
            for col in range(1, 9):
                ws.cell(r, col).fill = light
        # Wrap message column
        ws.cell(r, 6).alignment = Alignment(wrap_text=True)
        ws.row_dimensions[r].height = 40

    # Column widths
    col_widths = [5, 22, 28, 16, 28, 55, 20, 10]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze header
    ws.freeze_panes = 'A3'

    # Save to response
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="spk_contacts.xlsx"'
    return response


# ─────────────────────────────────────────────────────────────
# ORIGINAL PUBLIC VIEWS (unchanged)
# ─────────────────────────────────────────────────────────────
def home(request):
    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        email   = request.POST.get('email', '').strip()
        phone   = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        if name and email and message:
            Contact.objects.create(name=name, email=email, phone=phone,
                                   subject=subject, message=message)
            messages.success(request, 'Thank you! We will get back to you within 24 hours.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('home')

    events  = Event.objects.filter(is_active=True).order_by('order', '-created_at')
    gallery = Gallery.objects.filter(is_active=True).order_by('order', '-created_at')
    return render(request, 'home.html', {'events': events, 'gallery': gallery})


def about(request):
    return render(request, 'about.html')

def events_page(request):
    events = Event.objects.filter(is_active=True).order_by('order', '-created_at')
    return render(request, 'events.html', {'events': events})

def gallery_page(request):
    gallery = Gallery.objects.filter(is_active=True).order_by('order', '-created_at')
    return render(request, 'gallery.html', {'gallery': gallery})

def contact(request):
    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        email   = request.POST.get('email', '').strip()
        phone   = request.POST.get('phone', '').strip()
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        if name and email and message:
            Contact.objects.create(name=name, email=email, phone=phone,
                                   subject=subject, message=message)
            messages.success(request, 'Message sent! We will reply within 24 hours.')
        else:
            messages.error(request, 'Please fill in all required fields.')
        return redirect('contact')
    return render(request, 'contact.html')
